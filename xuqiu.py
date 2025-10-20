import ctypes
import logging
import os
import stat
import time
import datetime
import re
from asyncio import sleep
from concurrent.futures import ThreadPoolExecutor
from clinet.xuqiu_ui import Ui_Form
from clinet.create_xq_ui import Ui_create_xq
from PySide6.QtWidgets import QDateEdit
from PySide6.QtCore import QDate
from PySide6.QtGui import QIntValidator
import pyperclip
from PySide6 import QtWidgets, QtCore
from PySide6.QtCore import QTime, QDateTime, QDir, QAbstractTableModel, QModelIndex
from PySide6.QtGui import Qt, QIcon, QKeySequence, QStandardItemModel, QColor, QAction, QShortcut
from PySide6.QtWidgets import QTextEdit, QApplication, QMessageBox, QDateTimeEdit, QLineEdit, QPushButton, QTextBrowser, \
    QFormLayout, QHBoxLayout, QComboBox, QInputDialog, QMainWindow, QFileDialog, QTableWidget, QTableWidgetItem, \
    QHeaderView, QAbstractItemView,  QWidget, QTableView, QDialog, QVBoxLayout, \
    QDialogButtonBox, QLabel, QToolTip, QRadioButton, QMenu
from PySide6.QtUiTools import QUiLoader
from PySide6.QtGui import QStandardItem
from model.model import Project_detail, SQL_arg, History, dp, create_tables, ARG_model, db, XQ_INFO, JB_INFO, YSJ_KJ, \
    XQ_TABLE_INFO, TABLE_INFO, TABLE_COLUM_INFO, MD_INFO
from sql_helper.helper_restruct import Reader_Factory, Genner_Com
from collections import deque
import shutil
import subprocess
from conf import XUQIU_MOBAN_PATH, VSDX_PATH, GONGZUOQU_PATH, DATA_PATH, HSZ_PATH, VS_PATH, WPS_PATH, EDGE_PATH, \
    LSJS_PATH, XQLX, \
    TXT_PATH, ZDXQ, ZIP_PATH, MD_PATH
from sql_helper.read_sql_file import Reader_SQL
from tijiaobanben import Submit_banben_UI
from yuanshuju import CJJS_UI, BGL_UI

XUQIU_MOBAN_PATH = XUQIU_MOBAN_PATH
GONGZUOQU_PATH=GONGZUOQU_PATH
DATA_PATH = DATA_PATH
HSZ_PATH = HSZ_PATH
VS_PATH=VS_PATH
VSDX_PATH=VSDX_PATH
WPS_PATH=WPS_PATH
EDGE_PATH=EDGE_PATH
ZDXQ = ZDXQ
MD_PATH = MD_PATH
print(GONGZUOQU_PATH)
# XUQIU_TTYPE= ['任务','取数']
XUQIU_TTYPE = XQLX.split(',')

def get_files_in_directory(directory):
    files = []
    dirs = [] #文件夹
    try:
        for filename in os.listdir(directory):
            if os.path.isfile(os.path.join(directory, filename)):
                files.append(filename)
            else:
                if filename=='代码':
                    continue
                if filename.endswith('.assets'):
                    continue
                dirs.append(filename)
        return files,dirs
    except Exception as e:
        raise Exception("Failed to execute script 'main' due to unhandled exception!")
class Create_XUQIU_UI(QWidget):
    def __init__(self,wb):
        super().__init__()
        self.ui = Ui_create_xq()
        self.ui.setupUi(self)
        self.setStyleSheet("""
            QWidget {
                background-color: #e0f7fa;
                color: #333;
            }
            QPushButton {
                background-color: #0288d1;
                color: white;
                border: none;
                padding: 5px 10px;
                text-align: center;
                font-size: 16px;
                margin: 4px 2px;
            }
            QPushButton:hover {
                background-color: white;
                color: black;
                border: 2px solid #0288d1;
            }
            QLabel {
                font-size: 14px;
                color: #333;
            }
            QLineEdit, QComboBox {
                background-color: white;
                color: #333;
                border: 1px solid #ccc;
                padding: 5px;
            }
        """)
        self.tl_xqmc:QLineEdit = self.ui.lineEdit
        self.com_xqlx:QComboBox = self.ui.comboBox
        self.tl_tcr:QLineEdit = self.ui.lineEdit_2
        self.com_srmm:QComboBox=self.ui.comboBox_2
        self.com_scmm:QComboBox = self.ui.comboBox_3
        self.bt_qd :QPushButton =self.ui.pushButton
        self.bt_qd.clicked.connect(self.bcxq)
        self.wb = wb

        self.da_start_date:QDateEdit = self.ui.da_start_date
        self.da_end_date:QDateEdit = self.ui.da_end_date
        self.li_rentian:QLineEdit = self.ui.li_rentian
        # 开始时间设置当前日期
        self.da_start_date.setDate(QDate.currentDate())
        # 启用日历弹出
        self.da_start_date.setCalendarPopup(True)
        self.da_start_date.setDisplayFormat('yyyy-MM-dd')
        self.da_start_date.setDateRange(QDate(2020, 1, 1), QDate(2050, 12, 31))
        # 限制 li_rentian 只能输入数字
        self.li_rentian.setValidator(QIntValidator())
        self.li_rentian.setText('0')
        # 根据人天，计算结束时间
        self.li_rentian.textChanged.connect(self.f_li_rentian)
        # 设置结束时间
        self.da_end_date.setDate(QDate.currentDate())
        # 启用结束日期的日历弹出
        self.da_end_date.setCalendarPopup(True)
        self.da_end_date.setDisplayFormat('yyyy-MM-dd')
        self.da_end_date.setDateRange(QDate(2020, 1, 1), QDate(2050, 12, 31))


        self.inint_xqlx()
        self.init_com_srmm()

    def f_li_rentian(self):
        # 根据人天，计算结束时间
        rentian = self.li_rentian.text()
        if rentian == '':
            return
        try:
            rentian = int(rentian)
        except ValueError:
            QMessageBox.warning(self, '错误', '请输入有效的数字')
            return
        
        start_date = self.da_start_date.date()
        # 使用 QDate 的 addDays 方法进行日期计算
        end_date = start_date.addDays(rentian)
        self.da_end_date.setDate(end_date)


    def init_com_srmm(self):
        #输入模版 输出模版初始化
        files,_ = get_files_in_directory(XUQIU_MOBAN_PATH)
        for i in files:
            if i.startswith('输入'):
                self.com_srmm.addItem(i)
            elif i.startswith('输出'):
                self.com_scmm.addItem(i)
            else:
                continue


    def inint_xqlx(self):
        #需求类型
        self.com_xqlx.clear()
        for i in XUQIU_TTYPE:
            self.com_xqlx.addItem(i)

    # 创建md文档
    def create_md(self,name):
        # 创建md文档
        md_path = os.path.join(DATA_PATH,name,f'{name}.md')
        print(f'创建md文档: {md_path}')
        if os.path.exists(md_path):
            return
        with open(md_path,'w',encoding='utf-8') as f:
            f.write(f'# {name}')

    

    def init_xuqiu(self,name,srmb,scmb):
        #创建数据目录，如果目录已存在，则报错
        pt = os.path.join(DATA_PATH,name)
        srpt = os.path.join(XUQIU_MOBAN_PATH,srmb)
        srhz = srmb.split('.')[-1]
        scpt=os.path.join(XUQIU_MOBAN_PATH,scmb)
        schz = scmb.split('.')[-1]
        if os.path.exists(pt):
            raise ValueError('目录已存在，请勿重复创建')
        else:
            inint_path_str =['数据','文档']
            inint_path =[]
            for i in inint_path_str:
                inint_path.append(os.path.join(pt,i))
            os.mkdir(pt)
            for i in inint_path:
                os.mkdir(i)
            os.mkdir(os.path.join(pt,'代码'))
            shutil.copy(srpt,os.path.join(pt,f'输入-{name}.{srhz}'))
            shutil.copy(scpt,os.path.join(pt, f'输出-{name}.{schz}'))
            # 创建需求定义md文档
            self.create_md(name)
        return pt


    def bcxq(self):
        # 确定，保存需求
        try:
            name = self.tl_xqmc.text()
            srmb = self.com_srmm.currentText()
            scmb = self.com_scmm.currentText()
            path = self.init_xuqiu(name,srmb,scmb)
            xuqiu_type = self.com_xqlx.currentText()
            start_date = self.da_start_date.date().toString('yyyy-MM-dd')
            end_date = self.da_end_date.date().toString('yyyy-MM-dd')
            rentian = self.li_rentian.text()
            if rentian == '':
                rentian = 0
            else:
                rentian = int(rentian)
            status='开发中'
            tcr = self.tl_tcr.text()
            ret_dic = {'name':name,'path':path,'xuqiu_type':xuqiu_type,'status':status,'tcr':tcr
                       ,'start_date':start_date,'end_date':end_date,'rentian':rentian}
            xq_info = XQ_INFO(**ret_dic)
            xq_info.save()
            self.wb.init_tb_xuqiu()
            self.close()
        except Exception as e:
            QMessageBox.warning(self, '', e.__str__())



xuqiu_header = ['id','需求名称','需求类型','需求状态','需求提出人','需求存储位置','创建日期','开始时间','结束时间','总天数','剩余天数']
xuqiu_header_id =0
xuqiu_header_xqmc =1
xuqiu_header_xqlx =2
xuqiu_header_xqzt =3
xuqiu_header_xqtcr =4
xuqiu_header_xqccwz =5
xuqiu_header_cjrq =6
xuqiu_header_kssj =7
xuqiu_header_jzsj =8
xuqiu_header_rtr =9
xuqiu_header_srtr =10

xuqiu_dtl_header = ['id','类型','文件名','版本','备注','创建日期','更新日期','路径']
xuqiu_dtl_header_id=0
xuqiu_dtl_header_lx=1
xuqiu_dtl_header_wjm=2
xuqiu_dtl_header_bb=3
xuqiu_dtl_header_bz=4
xuqiu_dtl_ctime = 5
xuqiu_dtl_mtime = 6
xuqiu_dtl_header_lj=7
mhsheet="""
        QTableView {
            background-color: #f2f2f2;
            border: 1px solid #ccc;
            font-family: Arial;
            font-size: 16px;
        }
        
        QTableView::item {
            padding: 5px;
            border: none;
        }
        
        QTableView::item:selected {
            background-color: #b3d9ff;
        }
"""

jiaoben_dtl_header =['脚本ID','脚本名称','脚本备注','原需求id','原需求名称']
jiaoben_dtl_header_id = 0
jiaoben_dtl_header_name = 1
jiaoben_dtl_header_bz = 2
jiaoben_dtl_header_yxqid = 3
jiaoben_dtl_header_yxqmc = 4

class MyTableView(QTableView):
    def __init__(self, parent=None,vv=None):
        super().__init__(vv)
        self.setMouseTracking(True)
        self.setEditTriggers(QAbstractItemView.NoEditTriggers)

    def mouseMoveEvent(self, event):
        index = self.indexAt(event.pos())
        if index.isValid():
            item = self.model().itemFromIndex(index)
            if item:
                QToolTip.showText(event.globalPos(), item.text())


from clinet.XQXZ_UI import Ui_XQXZ
# 需求选择
class XQXZ_UI(QWidget):
    def __init__(self,xqdtl_id,parent):
        super().__init__()
        self.ui = Ui_XQXZ()
        self.ui.setupUi(self)
        self.init_ui()
        self.xqdtl_id = xqdtl_id
        self.parent = parent
    
    def init_ui(self):
        self.com_input:QComboBox = self.ui.com_input
        self.bt_ok:QPushButton = self.ui.pb_submit
        self.bt_ok.clicked.connect(self.f_bt_ok)
        self.li_input:QLineEdit = self.ui.li_input
        self.li_input.textChanged.connect(self.f_li_input)


    def f_li_input(self):
        # li 变更时，更新 com_input 下拉框内容，如果 li_input 为空，则 com_input 为空，
        # XQ_INFO 表中，所有需求名称
        xqmc = self.li_input.text()
        xq_infos = XQ_INFO.select().where(XQ_INFO.name.contains(xqmc))
        if len(xq_infos) == 0:
            self.com_input.clear()
        else:
            self.com_input.clear()
            for xq_info in xq_infos:
                # item = f'{xq_info.id}$${xq_info.name}'
                self.com_input.addItem(xq_info.name,xq_info.id)
    
    def f_bt_ok(self):
        # 返回需求名称，需求id
        xq_info = self.com_input.currentText()
        xq_id = self.com_input.currentData()
        xq_info = XQ_INFO.get_by_id(xq_id)
        xq_path = xq_info.path
        
        # 移动脚本到指定的需求中
        jb_info = JB_INFO.get_by_id(self.xqdtl_id)
        old_xq_id = jb_info.xq_id
        jb_old_path = jb_info.path
        jb_new_path = os.path.join(xq_path,jb_info.name)

        # 判断旧路径文件是否存在
        if not os.path.exists(jb_old_path):
            QMessageBox.warning(self,'提醒','脚本文件不存在')
            return
        # 判断新路径目录是否存在 xq_path
        if not os.path.exists(xq_path):
            QMessageBox.warning(self,'提醒','需求目录不存在')
            return
        # 判断新路径下是否有同名文件
        if os.path.exists(jb_new_path):
            QMessageBox.warning(self,'提醒','新路径下已有同名文件')
            return
        # 复制文件到新路径
        shutil.copy(jb_old_path,jb_new_path)
        # 更新数据库中内容
        jb_info.path = jb_new_path
        jb_info.xq_id = xq_id
        jb_info.save()
        # 删除旧文件
        os.remove(jb_old_path)
        self.parent.inint_tb_xuqiudtl(old_xq_id)
        self.close()


from clinet.JBXZ_UI import Ui_XQXZ as Ui_JBXZ
from model.model import YY_JB_INFO
# 脚本选择
class JBXZ_UI(QWidget):
    def __init__(self,xq_id,parent):
        super().__init__()
        self.ui = Ui_JBXZ()
        self.ui.setupUi(self)
        
        self.xq_id = xq_id # 目标需求id 
        self.parent = parent

        self.com_input:QComboBox = self.ui.com_input
        self.com_input.currentTextChanged.connect(self.f_com_input)
        self.bt_ok:QPushButton = self.ui.pb_submit
        self.bt_ok.clicked.connect(self.f_bt_ok)
        self.li_input:QLineEdit = self.ui.li_input
        self.li_input.textChanged.connect(self.f_li_input)
        # tb_jiaoben 显示列名为 脚本id，脚本名称，脚本备注
        
        self.tb_jiaoben:QTableView = self.ui.tb_jiaoben
        self.tb_jiaoben.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.tb_jiaoben.horizontalHeader().setStretchLastSection(True)
        
        self.tb_jiaoben.setSelectionBehavior(QAbstractItemView.SelectRows)
        # INSERT_YOUR_CODE
        # 设置选中行为的背景颜色为深蓝色
        # 通过设置QTableView的样式表来实现
        self.tb_jiaoben.setStyleSheet("QTableView::item:selected { background-color: #00008B; }")

        model = QStandardItemModel()
        for column, name in enumerate(jiaoben_dtl_header):
            header_item = QStandardItem(name)
            model.setHorizontalHeaderItem(column, header_item)
        self.tb_jiaoben.setModel(model)

    def get_row_content(self,index):
        try:
        # 获取行内容
            row = [index.model().data(index.sibling(index.row(), column), Qt.DisplayRole) for column in
                   range(index.model().columnCount())]
            return row
        except Exception as e:
            raise ValueError('未正确选择行')
    def f_li_input(self):
        # li 变更时，更新 com_input 下拉框内容，如果 li_input 为空，则 com_input 为空，
        # XQ_INFO 表中，所有需求名称
        xqmc = self.li_input.text()
        xq_infos = XQ_INFO.select().where(XQ_INFO.name.contains(xqmc))
        if len(xq_infos) == 0:
            self.com_input.clear()
        else:
            self.com_input.clear()
            for xq_info in xq_infos:
                # item = f'{xq_info.id}$${xq_info.name}'
                self.com_input.addItem(xq_info.name,xq_info.id)

    # COM 变化时，更新选择的脚本
    def f_com_input(self):
        xq_id = self.com_input.currentData()
        xq_info = XQ_INFO.get_by_id(xq_id)
        xq_name = xq_info.name
        jbs = JB_INFO.select().where(JB_INFO.xq_id == xq_id).where(JB_INFO.type == '脚本' and JB_INFO.version == '0')
        if len(jbs) == 0:
            self.tb_jiaoben.clear()
        else:
            model = QStandardItemModel()
            for column, name in enumerate(jiaoben_dtl_header):
                header_item = QStandardItem(name)
                model.setHorizontalHeaderItem(column, header_item)
            for jb in jbs:
                jb_id = jb.id
                jb_name = jb.name
                jb_desc = jb.desc
                item = [str(jb_id),jb_name,jb_desc,str(xq_id),xq_name]
                row_items =[QStandardItem(item) for item in item]
                model.appendRow(row_items)
            self.tb_jiaoben.setModel(model)
    
    def genner_jiaoben_info(self,datas):
        ret = zip(jiaoben_dtl_header,datas)
        return dict(ret)



    # 获取选中脚本的信息
    def get_select_jiaoben_info(self):
        # 获取选中脚本的信息
        select_idxs = self.tb_jiaoben.selectedIndexes()
        if not select_idxs:
            # 如果没有选中任何项，返回空列表
            return []

        # 去重
        row_indexs = set()
        
        ret = []
        # 遍历所有选中的索引
        for idx in select_idxs:
            # 获取每个选中行的内容
            if idx.row() in row_indexs:
                continue
            row_indexs.add(idx.row())
            datas = self.get_row_content(idx)
            # 将行内容转换为字典格式
            tmp = self.genner_jiaoben_info(datas)
            # 将字典添加到返回列表中
            ret.append(tmp)

        # 返回所有选中行的信息
        return ret

    def f_bt_ok(self):
        jbs = self.get_select_jiaoben_info()
        # 引用的目标需求id
        target_xq_id = XQ_INFO.get_by_id(self.xq_id)
        for jb in jbs:
            jb_id = jb.get('脚本ID')
            jb_desc = jb.get('脚本备注')
            jb_yxqid = jb.get('原需求id')
            # 判断目标需求下是否已经有了引用，有了就跳过
            yy_jbs = YY_JB_INFO.select().where(YY_JB_INFO.t_xq_id==target_xq_id).where(YY_JB_INFO.sjb_id==jb_id)
            if len(yy_jbs) > 0:
                continue
            # 创建一条新的引用脚本信息
            yy_jb_info = YY_JB_INFO(
                s_xq_id=jb_yxqid,
                sjb_id=jb_id,
                t_xq_id=target_xq_id,
                yy_desc=jb_desc
            )
            yy_jb_info.save()
            # 复制一条新的脚本信息
        # 返回所有选中行的信息
        self.parent.inint_tb_xuqiudtl(target_xq_id)
        self.close()


class XUQIU_UI(QWidget):
    def __init__(self,main_pk):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)
        self.tb_xuqiu:QTableView=self.ui.tb_xuqiu
        self.main_pk = main_pk

        self.tb_xuqiu.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.bt_rename:QPushButton = self.ui.bt_rename
        self.bt_rename.clicked.connect(self.f_bt_rename)

        self.bt_create:QPushButton = self.ui.bt_create
        self.bt_create.clicked.connect(self.create_xuqiu)
        self.bt_remove:QPushButton=self.ui.bt_remove
        self.bt_remove.clicked.connect(self.remove_xuqiu)
        self.li_search:QLineEdit=self.ui.lineEdit

        model = QStandardItemModel()



        self.li_search.textChanged.connect(self.search_xuqiu)
        self.li_search.selectionChanged.connect(self.select_xuqiu)


        self.com_type:QComboBox = self.ui.com_type
        self.com_type.addItems(['','文档','脚本','文件夹'])
        self.com_type.currentTextChanged.connect(self.f_com_type)
        self.tb_xuqiudtl:QTableView =self.ui.tb_xuqiudtl
        self.tb_xuqiudtl.setSortingEnabled(True)
        self.tb_xuqiudtl.resizeColumnsToContents()
        self.tb_xuqiu.setSortingEnabled(True)
        self.tb_xuqiu.resizeColumnsToContents()


        self.com_sear_type:QComboBox=self.ui.com_sear_type
        self.com_sear_type.currentTextChanged.connect(self.init_tb_xuqiu)


        self.tb_xuqiudtl.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.tb_xuqiudtl.setWordWrap(True) 

        self.bt_new_jiaoben:QPushButton =self.ui.new_jiaoben
        self.bt_new_jiaoben.clicked.connect(self.new_jiaoben)
        self.bt_submit_jiaoben:QPushButton=self.ui.submit_jiaoben
        self.bt_submit_jiaoben.clicked.connect(self.submit_jiaoben)
        self.tb_xuqiudtl.doubleClicked.connect(self.op_wenjian)
        self.bt_rm_jb:QPushButton=self.ui.bt_rm_jb
        self.bt_rm_jb.clicked.connect(self.rm_jb)
        self.tb_xuqiu.setStyleSheet(mhsheet)

        for column, name in enumerate(xuqiu_header):
            header_item = QStandardItem(name)
            model.setHorizontalHeaderItem(column, header_item)
        self.tb_xuqiu.setModel(model)
        self.bt_daoru:QPushButton =self.ui.bt_daoru
        self.bt_daoru.clicked.connect(self.f_bt_daoru)

        self.bt_xqzt:QPushButton = self.ui.bt_xqzt
        self.bt_xqzt.clicked.connect(self.f_bt_xqzt)
        self.ra_banben:QRadioButton = self.ui.ra_banben

        # 提交md文档
        self.bt_tj_md:QPushButton = self.ui.bt_tj_md
        self.bt_tj_md.clicked.connect(self.f_bt_tj_md)

        self.tb_xuqiu.clicked.connect(self.dj_xuqiu)
        self.tb_xuqiu.doubleClicked.connect(self.ddj_xuqiu)
        self.li_search_wd:QLineEdit=self.ui.li_search_wd
        self.li_search_wd.textChanged.connect(self.f_com_type)
        self.init_tb_xuqiu()
        self.tb_xuqiu.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.tb_xuqiu.horizontalHeader().setStretchLastSection(True)
        self.tb_xuqiudtl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.tb_xuqiudtl.horizontalHeader().setStretchLastSection(True)
        self.ra_banben.clicked.connect(self.f_com_type)

        self.bt_cjjs:QPushButton = self.ui.bt_cjjs
        self.bt_cjjs.clicked.connect(self.f_cjjs)

        self.read_sql_eng = Reader_SQL()
        self.com_sear_type.addItems(['','开发中','完成'])

        self.is_zd: QRadioButton = self.ui.is_zd
        self.is_zd.toggled.connect(self.f_is_zd)
        # self.is_zd.setChecked(True)
        shortcut2 = QShortcut(QKeySequence(Qt.CTRL | Qt.Key_B), self)
        shortcut2.activated.connect(self.f_bgl)

        shortcut3 = QShortcut(QKeySequence(Qt.CTRL | Qt.Key_Z), self)
        shortcut3.activated.connect(self.f_tz_parent)
        self.bgl_ui=None
        self.dict_search_tmp={}


        self.tb_xuqiu.hideColumn(5)

        self.is_guidang:QRadioButton = self.ui.is_guidang
        self.is_guidang.toggled.connect(self.f_com_type)

        self.la_xq_info:QLabel = self.ui.label_2

    # 设置当前需求名称
    def f_set_la_xq_info(self,id):
        xq_info = XQ_INFO.get_by_id(id)
        self.la_xq_info.setText(f'当前需求:《{xq_info.name}》')


    def get_current_xqdtl_info(self):
        idx = self.tb_xuqiudtl.currentIndex()
        datas = self.get_row_content(idx)
        ret = {
            'id':datas[xuqiu_dtl_header_id],
            '类型':datas[xuqiu_dtl_header_lx],
            '文件名':datas[xuqiu_dtl_header_wjm],
            '版本':datas[xuqiu_dtl_header_bb],
            '备注':datas[xuqiu_dtl_header_bz],
            '创建日期':datas[xuqiu_dtl_ctime],
            '更新日期':datas[xuqiu_dtl_mtime],
            '路径':datas[xuqiu_dtl_header_lj]
        }
        return ret

    # 获取选中的dtl数据
    def get_select_xqdtl_info(self):
        select_idxs = self.tb_xuqiudtl.selectedIndexes()
        if len(select_idxs) == 0:
            return None
        idx = select_idxs[0]
        ret = []
        for idx in select_idxs:
            datas = self.get_row_content(idx)
            tmp = {
                'id':datas[xuqiu_dtl_header_id],
                '类型':datas[xuqiu_dtl_header_lx],
                '文件名':datas[xuqiu_dtl_header_wjm],
                '版本':datas[xuqiu_dtl_header_bb],
                '备注':datas[xuqiu_dtl_header_bz],
            }
            ret.append(tmp)
        return ret

    def f_bt_tj_md(self):
        xqdtl_info = self.get_current_xqdtl_info()
        xq_dtl_id = xqdtl_info.get('id')
        xq_dtl_name = xqdtl_info.get('文件名')
        xq_info = self.get_current_xq_info()
        xq_id = xq_info.get('id')
        path = xqdtl_info.get('路径')
        xq_name = xq_info.get('需求名称')

        if not xq_dtl_name.endswith('.md'):
            QMessageBox.warning(self,'提醒','文件必须为md格式')
            return
        if xq_dtl_id =='-888':
            QMessageBox.warning(self,'提醒','请提交原始md文档,不可提交链接文档')
            return
        elif xq_dtl_id =='-999': # 首次提交的md文档
            tj_time = datetime.datetime.now()
            md_info = MD_INFO(xq_id=xq_id,name=xq_dtl_name,path=path,bz='',tj_time=tj_time,xq_name=xq_name)
            md_info.save()
            xq_item = XQ_INFO.get_by_id(xq_id)
            self.inint_tb_xuqiudtl(xq_item)
            file_id = md_info.file_id
            self.f_md_tijiao_zsk(file_id)
            self.repaint()
        else:
            md_info = MD_INFO.get_by_id(xq_dtl_id)
            md_info.tj_time = datetime.datetime.now()
            md_info.xq_name = xq_name
            md_info.save()
            xq_item = XQ_INFO.get_by_id(xq_id)
            self.inint_tb_xuqiudtl(xq_item)
            file_id = md_info.file_id
            self.f_md_tijiao_zsk(file_id)
            self.repaint()
    
    def f_md_tijiao_zsk(self,id):
        ## 传入文档id,通过调用AI接口,对Md文档存储到知识库中，进行检索增强
        print('传入文档id,通过调用AI接口,对Md文档存储到知识库中，进行检索增强',id)
        
        return {'status':True,'msg':'提交成功'}

    def f_tz_parent(self):
        if self.main_pk is None:
            QMessageBox.warning(self, '提醒', '没有上一级')
        else:
            self.main_pk.is_zd.setChecked(True)
            self.main_pk.is_zd.setChecked(self.is_zd.isChecked())
            self.main_pk.showNormal()



    def f_bgl(self):
        if self.bgl_ui is None:
            self.bgl_ui = BGL_UI()

        self.bgl_ui.is_zd.setChecked(True)
        self.bgl_ui.is_zd.setChecked(self.is_zd.isChecked())
        self.bgl_ui.hide()
        self.bgl_ui.show()

    # 超级检索
    def f_cjjs(self):
        select_idxs = self.tb_xuqiu.selectedIndexes()
        xqids = []
        for idx in select_idxs:
            datas = self.get_row_content(idx)
            item = datas[xuqiu_header_id]
            xqids.append(item)
        self.ysjgl_ui = CJJS_UI(xqids)
        self.ysjgl_ui.show()

    def contextMenuEvent(self, event):
        # 创建右键菜单
        menu = QMenu(self)
        # 创建QAction
        action = QAction("修改备注", self)
        action.triggered.connect(self.xgbz)

        action2 = QAction("复制文件", self)
        action2.triggered.connect(self.fzwj)

        action3 = QAction("归档", self)
        action3.triggered.connect(self.f_guidang)

        # 代码移动
        action4 = QAction("代码移动", self)
        action4.triggered.connect(self.f_code_move)

        # 脚本重命名
        action5 = QAction("脚本重命名", self)
        action5.triggered.connect(self.f_script_rename)

        # 脚本引用
        action6 = QAction("脚本引用", self)
        action6.triggered.connect(self.f_jiaobenyy)

        # 跳转到指定需求
        action7 = QAction("跳转到源", self)
        action7.triggered.connect(self.f_tiao_zhuan_xq)

        # 将QAction添加到菜单中
        menu.addAction(action)
        menu.addAction(action2)
        menu.addAction(action3)
        menu.addAction(action4)
        menu.addAction(action5)
        menu.addAction(action6)
        menu.addAction(action7)
        #
        # action2 = QAction("备注", self)
        # action2.triggered.connect(self.f_bzb)
        # # 将QAction添加到菜单中
        # menu.addAction(action2)
        # 显示菜单
        menu.exec_(event.globalPos())

    def f_code_move(self):
        # 选中要移动的代码，弹出对话框，输入要移动到的需求名称，将代码移动到指定的需求中
        cur_xq_dtl_info = self.get_current_xqdtl_info()

        if cur_xq_dtl_info.get('类型') != '脚本':
            QMessageBox.warning(self,'提醒','只能移动脚本')
            return
        cur_xqdtl_id = cur_xq_dtl_info.get('id')
        # 弹出一个对话框，对话框中包含一个输入框，输入框中包含一个下拉框，下拉框中包含所有需求名称
        
        self.xqxz_ui = XQXZ_UI(cur_xqdtl_id,self)
        self.xqxz_ui.show()
    
    # 脚本引用
    def f_jiaobenyy(self):
        cur_xq_info = self.get_current_xq_info()
        xq_id = cur_xq_info['id']
        self.jbxz_ui = JBXZ_UI(xq_id,self)
        self.jbxz_ui.show()

    def f_script_rename(self):
        # 脚本重命名
        cur_xq_dtl_info = self.get_current_xqdtl_info()
        if cur_xq_dtl_info.get('类型') != '脚本':
            QMessageBox.warning(self,'提醒','只能重命名脚本')
            return
        cur_xqdtl_id = cur_xq_dtl_info.get('id')
        jb_info = JB_INFO.get_by_id(cur_xqdtl_id)
        jb_name = jb_info.name
        jb_path = jb_info.path

        xq_id = jb_info.xq_id
        xq_info = XQ_INFO.get_by_id(xq_id)
        xq_path = xq_info.path

        # 弹出对话框，输入新名称
        dialog = QDialog()
        dialog.setWindowTitle('脚本重命名')
        dialog.setFixedSize(600, 200)
        layout = QVBoxLayout()
        lb1 = QLabel()
        lb1.setText('新名称')   
        in_name = QLineEdit()
        in_name.setText(jb_name)
        layout.addWidget(lb1)
        layout.addWidget(in_name)
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(button_box)
        dialog.setLayout(layout)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        result = dialog.exec_()
        
        if result == QDialog.Accepted:
            new_name = in_name.text()
            # 判断新名称是否为空
            if new_name == '':
                QMessageBox.warning(self,'提醒','新名称不能为空')
                return
            # 拼接新路径
            new_path = os.path.join(xq_path,new_name)
            # 判断新路径是否存在
            if os.path.exists(new_path):
                QMessageBox.warning(self,'提醒','新路径已存在')
                return
            # 复制文件到新路径
            shutil.copy(jb_path,new_path)
            # 更新数据库
            jb_info.path = new_path
            jb_info.name = new_name
            jb_info.save()
            # 删除旧文件
            os.remove(jb_path)
            self.inint_tb_xuqiudtl(xq_id)
            self.repaint()

    # 跳转到原需求,跳转到源
    def f_tiao_zhuan_xq(self):
        cur_xq_dtl = self.get_current_xqdtl_info()
        leixin = cur_xq_dtl.get('类型')
        path = cur_xq_dtl.get('路径')
        if leixin == '脚本引用':
            xq_id = cur_xq_dtl.get('id')
            item = YY_JB_INFO.get_by_id(xq_id)
            xq_id = item.s_xq_id
            self.f_set_la_xq_info(xq_id)
            self.inint_tb_xuqiudtl(xq_id)
        elif leixin =='文档':
            if 'http' in path:
                QMessageBox.warning(self,'提醒','文档路径为链接，不支持跳转')
                return
            path = os.path.normpath(path)
            # 判断路径是否存在，不存在不跳转
            if not os.path.exists(path):
                QMessageBox.warning(self,'提醒','文档路径不存在，不支持跳转')
                return
            os.system(f'explorer /select,"{path}"')
        else:
            QMessageBox.warning(self,'提醒',f'{leixin}类型文件不支持跳转')
            return
        
        
    

    def f_guidang(self):
        # 弹出对话框输入归档内容
        dialog = QDialog()
        dialog.setWindowTitle('归档')
        dialog.setFixedSize(600, 200)
        layout = QVBoxLayout()
        lb1 = QLabel()
        lb1.setText('归档内容')
        in_guidang = QLineEdit()
        layout.addWidget(lb1)
        layout.addWidget(in_guidang)
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(button_box)
        dialog.setLayout(layout)
        # 连接按钮事件，点击确认或取消时关闭弹窗
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        # 显示弹窗并等待用户操作
        result = dialog.exec_()
        if result == QDialog.Accepted:
            guidang = in_guidang.text()
            # 输出info级日志，记录归档内容
            print(f'归档内容: {guidang}')

            print('获取选择的dtl数据@@@@@@@开始')
            datas = self.get_select_xqdtl_info()
            ids = [data.get('id') for data in datas if data.get('类型') =='脚本']
            if len(ids) > 0:
                
                # 根据id 获取到 JB_INFO 列表，批量更新他们的gd_name 字段
                print('ids',ids)
                jbs = JB_INFO.select().where(JB_INFO.id.in_(ids))
                for jb in jbs:
                    jb.gd_name = guidang
                    jb.save()
            print('获取选择的dtl数据@@@@@@@结束')


    #修改备注
    def xgbz(self):
        idx = self.tb_xuqiudtl.currentIndex()
        datas = self.get_row_content(idx)
        
        cur_xq_dtl = self.get_current_xqdtl_info()
        id = cur_xq_dtl.get('id')
        bz = cur_xq_dtl.get('备注')
        leixin = cur_xq_dtl.get('类型')

        ret = self.show_rename_dialog2(bz)
        if leixin == '脚本':
            item = JB_INFO.get_by_id(id)
            xq_id = item.xq_id
        elif leixin == '脚本引用':
            item = YY_JB_INFO.get_by_id(id)
            xq_id = item.t_xq_id
        else:
            return
        if ret.get('bz'):
            if leixin == '脚本':
                item.desc = ret.get('bz')
            elif leixin == '脚本引用':
                item.yy_desc = ret.get('bz')
            item.save()
        self.inint_tb_xuqiudtl(xq_id)


    # 复制文件到剪切版
    def fzwj(self):
        select_idxs = self.tb_xuqiudtl.selectedIndexes()
        path = []
        for idx in select_idxs:
            datas =self.get_row_content(idx)
            item = datas[xuqiu_dtl_header_lj]
            path.append(item)
        tmp = ','.join(path)
        str = f'Get-Item {tmp} | Set-Clipboard'
        args = ['powershell', str]
        subprocess.Popen(args =args)



    def show_rename_dialog2(self,bz):
        # 创建一个修改需求的对话框
        dialog = QDialog()
        dialog.setWindowFlag(QtCore.Qt.WindowStaysOnTopHint)
        dialog.setWindowTitle('修改备注')
        dialog.setFixedSize(600, 200)
        layout = QVBoxLayout()

        #备注信息
        lb1 = QLabel()
        lb1.setText('备注信息')
        in_bmss = QLineEdit()
        in_bmss.setText(bz)

        layout.addWidget(lb1)
        layout.addWidget(in_bmss)

        # 创建一个按钮框
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(button_box)
        # 设置按钮的点击事件
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        # 将布局设置给对话框
        dialog.setLayout(layout)
        # 显示对话框并等待用户点击按钮
        result = dialog.exec_()
        # 获取用户输入的文本
        if result == QDialog.Accepted:
            return { 'bz':in_bmss.text()}
        else:
            return {}



    def f_is_zd(self,checked):
        flags = self.windowFlags()
        if checked:
            self.setWindowFlags(flags | Qt.WindowStaysOnTopHint)
        else:
            self.setWindowFlags(flags & ~Qt.WindowStaysOnTopHint)
        self.show()


    def show_rename_dialog(self,xqlx,tcr,xqmc,kssj=None,rtr=None):
        # 创建一个修改需求的对话框
        dialog = QDialog()
        dialog.setWindowFlag(QtCore.Qt.WindowStaysOnTopHint)
        dialog.setWindowTitle('修改需求')
        dialog.setFixedSize(600,350)
        layout = QVBoxLayout()

        #需求名称
        lb2 = QLabel()
        lb2.setText('需求名称')
        in_xqmc = QLineEdit()
        in_xqmc.setText(xqmc)

        lb0 = QLabel()
        lb0.setText('需求类型')
        in_xqlx = QComboBox()
        in_xqlx.clear()
        in_xqlx.addItems(XUQIU_TTYPE)
        in_xqlx.setCurrentText(xqlx)
        # 创建一个输入框
        lb1 = QLabel()
        lb1.setText('提出人姓名')
        in_tcr = QLineEdit()
        in_tcr.setText(tcr)

        # 如果kssj 为空，则设置为当前日期
        if kssj is None:
            kssj = QDate.currentDate()
        else:
            # 时间格式为 2025-06-20 00:00:00 需要转换为 QDate
            kssj_str = kssj.strftime('%Y-%m-%d')
            # 使用转换后的字符串来创建QDate对象
            kssj = QDate.fromString(kssj_str, 'yyyy-MM-dd')
        if rtr is None:
            rtr = '0'
        else:
            rtr = str(rtr)

        # 开始时间
        lb5 = QLabel()
        lb5.setText('开始时间')
        in_kssj = QDateEdit()
        in_kssj.setDate(kssj)
        in_kssj.setCalendarPopup(True)
        in_kssj.setDisplayFormat('yyyy-MM-dd')

        # 人天
        lb4 = QLabel()
        lb4.setText('人天')
        in_rtr = QLineEdit()
        in_rtr.setText(rtr)
        in_rtr.setValidator(QIntValidator())

        



        layout.addWidget(lb2)
        layout.addWidget(in_xqmc)
        layout.addWidget(lb0)
        layout.addWidget(in_xqlx)
        layout.addWidget(lb1)
        layout.addWidget(in_tcr)
        layout.addWidget(lb5)
        layout.addWidget(in_kssj)

        layout.addWidget(lb4)
        layout.addWidget(in_rtr)


        # 创建一个按钮框
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(button_box)
        # 设置按钮的点击事件
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        # 将布局设置给对话框
        dialog.setLayout(layout)
        # 显示对话框并等待用户点击按钮
        result = dialog.exec_()
        # 获取用户输入的文本
        if result == QDialog.Accepted:
            return {'xqlx':in_xqlx.currentText(),'tcr':in_tcr.text(),'xqmc':in_xqmc.text(),'ksrq':in_kssj.date().toString('yyyy-MM-dd'),'rtr':in_rtr.text()}
        else:
            return {}
    def f_bt_rename(self):
        idx = self.tb_xuqiu.currentIndex()
        if idx.isValid():
            datas = self.get_row_content(idx)
            id = datas[0]
            itme = XQ_INFO.get_by_id(id)
            ret = self.show_rename_dialog(itme.xuqiu_type,itme.tcr,itme.name,itme.start_date,itme.rentian)
            # 根据人天，开始时间计算结束时间
            ksrq = ret.get('ksrq')
            print(f"ksrq:{ksrq}")
            rtr = ret.get('rtr')
            # 字符串转日期
            ksrq = QDate.fromString(ksrq, 'yyyy-MM-dd')
            jzrq = ksrq.addDays(int(rtr))
            itme.start_date = ksrq.toString('yyyy-MM-dd')
            itme.end_date = jzrq.toString('yyyy-MM-dd')
            itme.rentian = rtr
            itme.xuqiu_type = ret.get('xqlx')
            itme.tcr = ret.get('tcr')
            print(f"*******************{ ret.get('xqmc')}")
            itme.name = ret.get('xqmc')
            itme.save()
            self.init_tb_xuqiu()

    def f_bt_xqzt(self):
        idx = self.tb_xuqiu.currentIndex()
        if idx.isValid():
            datas = self.get_row_content(idx)
            id = datas[0]
            xq_info = XQ_INFO.get_by_id(id)
            if xq_info.status =='开发中':
                xq_info.status='完成'
            else:
                xq_info.status ='开发中'
            xq_info.save()
        self.init_tb_xuqiu()

# xuqiu_header = ['id','需求名称','需求类型','需求状态','需求提出人','需求存储位置','创建日期']
# xuqiu_header_id =0
# xuqiu_header_xqmc =1
# xuqiu_header_xqlx =2
# xuqiu_header_xqzt =3
# xuqiu_header_xqtcr =4
# xuqiu_header_xqccwz =5
# xuqiu_header_cjrq =6
    def get_current_xq_info(self):
        idx = self.tb_xuqiu.currentIndex()
        datas = self.get_row_content(idx)
        ret = {
            'id':datas[xuqiu_header_id],
            '需求名称':datas[xuqiu_header_xqmc],
            '需求类型':datas[xuqiu_header_xqlx],
            '需求状态':datas[xuqiu_header_xqzt],
            '需求提出人':datas[xuqiu_header_xqtcr],
            '需求存储位置':datas[xuqiu_header_xqccwz],
            '创建日期':datas[xuqiu_header_cjrq]  
        }
        return ret
    


    def f_bt_daoru(self):
        try:
            idx = self.tb_xuqiu.currentIndex()
            d_datas = self.get_row_content(idx)
            dpth = d_datas[xuqiu_header_xqccwz]
            # 导入文件
            initial_dir = QDir(LSJS_PATH)
            file_dialog = QFileDialog()
            file_dialog.setDirectory(initial_dir)

            file_dialog.setFileMode(QFileDialog.ExistingFiles)
            file_dialog.setNameFilter("All files (*.*);;doc files (*.docx);;xls file (*.xlsx)")
            if file_dialog.exec_():
                file_paths = file_dialog.selectedFiles()
                for file_path in file_paths:
                    s_pth = file_path
                    shutil.copy(s_pth,dpth)
            self.f_com_type()
        except Exception as e:
            QMessageBox.warning(self,'',e.__str__())

# datas = self.get_row_content(index)
# path = datas[-1]
# os.startfile(path)
    def op_wenjian(self):
        idx = self.tb_xuqiudtl.currentIndex()
        datas = self.get_row_content(idx)
        print(datas[4])

        file_path = datas[-1]
        # file_path = file_path.replace(' ','')

        if datas[1]=='文档':
            if file_path.endswith('.pdf'):
                subprocess.Popen(F'{EDGE_PATH} "{file_path}"')
            elif file_path.endswith('.txt') or file_path.endswith('.sql'):
                subprocess.Popen(f'{TXT_PATH} "{file_path}"')
            elif file_path.endswith('.txt') or file_path.endswith('.vsdx'):
                subprocess.Popen(f'{VSDX_PATH} "{file_path}"')
                # VSDX_PATH
            elif file_path.endswith('.zip'):
                subprocess.Popen(f'{ZIP_PATH} "{file_path}"')
                # 打开zip文件
            elif file_path.startswith('http'):
                subprocess.Popen(F'{EDGE_PATH} "{file_path}"')
                pass
            elif file_path.endswith('.md'):
                subprocess.Popen(F'{MD_PATH}  "{file_path}"')
            elif file_path.endswith('.py'):
                subprocess.Popen(F'{VS_PATH} "{file_path}"')
            else:
                subprocess.Popen(F'{WPS_PATH} "{file_path}"')
        elif datas[1]=='文件夹':
            path = file_path
            os.startfile(path)
        elif datas[1]=='MD':
            subprocess.Popen(F'{MD_PATH}  "{file_path}"')
        elif datas[1]=='归档':
            pass
        else:
            print( F'{VS_PATH}  "{file_path}"')
            logging.info(f"尝试打开文件: {file_path}")
            # subprocess.Popen([VS_PATH, '--goto', datas[-1].strip()])
            subprocess.Popen(F'{VS_PATH}  "{file_path}"')

    def set_file_permissions(self,file_path): #设置为可读可写
        try:
            # 获取文件属性
            file_attributes = ctypes.windll.kernel32.GetFileAttributesW(file_path)

            # 移除只读属性
            new_attributes = file_attributes & ~stat.FILE_ATTRIBUTE_READONLY

            # 设置新的文件属性
            ctypes.windll.kernel32.SetFileAttributesW(file_path, new_attributes)
            print(f"成功设置文件 {file_path} 为可读可写")
        except Exception as e:
            print(f"设置文件权限失败：{e}")
    def rm_jb(self):

        cur_xq_dtl = self.get_current_xqdtl_info()
        dtl_leixin = cur_xq_dtl.get('类型')
        if dtl_leixin == '脚本引用':
            out_str = '是否删除"引用"'
        elif dtl_leixin == '脚本':
            out_str = '是否删除"脚本"'
        else:
            out_str = '是否删除选中内容'

        #删除脚本
        result = QMessageBox.question(self, "提醒", out_str, QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if result == QMessageBox.No:
            return
        
        
        if dtl_leixin == '脚本':
            id = int(cur_xq_dtl.get('id'))
            # 删除脚本
            if id>=0:
                data = JB_INFO.get_by_id(id)
                path = data.path
                if os.path.exists(path):
                    self.set_file_permissions(path)
                    os.remove(path)
                tmp = JB_INFO.get_by_id(id)
                kjs = YSJ_KJ.select().where(YSJ_KJ.jb_id == tmp)
                for i in kjs:
                    YSJ_KJ.delete_by_id(i.id)
                JB_INFO.delete_by_id(id)
                xqis = XQ_TABLE_INFO.select().where(XQ_TABLE_INFO.jb_id==tmp)
                for i in xqis:
                    XQ_TABLE_INFO.delete_by_id(i.id)
                
                # 删除引用中的脚本记录
                yy_jbs = YY_JB_INFO.select().where(YY_JB_INFO.sjb_id==tmp)
                for i in yy_jbs:
                    YY_JB_INFO.delete_by_id(i.id)
                self.f_com_type()
        elif dtl_leixin == '文档':
            QMessageBox.warning(self,'','文档请去对应目录自行删除')
        elif dtl_leixin == '脚本引用':
            id = int(cur_xq_dtl.get('id'))
            if id>=0:
                data = YY_JB_INFO.get_by_id(id)
                data.delete_by_id(id)
                self.f_com_type()
        else:
            QMessageBox.warning(self,'','文档请去对应目录自行删除')



        


    def show_input_dialog_submit(self):
        # 创建一个对话框
        dialog = QDialog()
        dialog.setWindowTitle('提交脚本')
        dialog.setFixedSize(600,200)
        layout = QVBoxLayout()
        lb1 = QLabel()
        lb1.setText('备注')
        in_bz = QLineEdit()
        layout.addWidget(lb1)
        layout.addWidget(in_bz)
        # 创建一个按钮框
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(button_box)
        # 设置按钮的点击事件
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        # 将布局设置给对话框
        dialog.setLayout(layout)
        # 显示对话框并等待用户点击按钮
        result = dialog.exec_()
        # 获取用户输入的文本
        if result == QDialog.Accepted:
            return {'bz':in_bz.text()}
        else:
            return {'bz':''}

    def submit_jiaoben(self):
        #提交版本
        try:
            index = self.tb_xuqiudtl.currentIndex()
            if index.isValid():
                index_datas = self.get_row_content(index)
                idx = int(index_datas[xuqiu_dtl_header_id])
                version = index_datas[xuqiu_dtl_header_bb]
                print(idx,version)
                if idx < 0 or version!='0':
                    raise ValueError('只可提交工作区代码（版本为0）')
                self.tijiaobanben_ui = Submit_banben_UI(idx)
                self.tijiaobanben_ui.show()

        except Exception as e:
            print(e.__str__())
            QMessageBox.warning(self,'',e.__str__())


    def new_jiaoben(self):
        # 新建脚本
        try:
            index = self.tb_xuqiu.currentIndex()
            if not index.isValid():
                return
            data = self.get_row_content(index)
            id = int(data[0])
            xq_id = XQ_INFO.get_by_id(id)
            diago = self.show_input_dialog(xq_id.name)
            if  diago.get('name') and diago.get('name').endswith('.sql'):
                pt =os.path.join(xq_id.path,diago.get('name'))
                if os.path.exists(pt):
                    QMessageBox.warning(self, '', '文件已存在，导入文件')
                else:
                    with open(pt,'w',encoding='utf-8') as f:
                        f.write(f"-- {diago.get('name')}::{diago.get('bz')}")
                name = diago.get('name')
                desc = diago.get('bz')
                type='脚本'
                version=0
                jb_info = JB_INFO()
                jb_info.xq_id = xq_id
                jb_info.name=name
                jb_info.desc=desc
                jb_info.type=type
                jb_info.version=0
                jb_info.path= pt
                jb_info.save()
                self.f_com_type()
            else:
                QMessageBox.warning(self,'','文件名为空或文件不合法')
                return
        except Exception as e:
            QMessageBox.warning(self,'',e.__str__())

    def show_input_dialog(self,name):
        # 创建一个对话框
        dialog = QDialog()
        dialog.setWindowFlags(dialog.windowFlags() | Qt.WindowStaysOnTopHint)
        dialog.setWindowTitle('新建脚本')
        dialog.setFixedSize(600,200)
        layout = QVBoxLayout()

        lb0 = QLabel()
        lb0.setText('文件名称')
        in_name = QLineEdit()
        in_name.setText(f"{name}.sql")
        # 创建一个输入框
        lb1 = QLabel()
        lb1.setText('备注')
        in_bz = QLineEdit()
        layout.addWidget(lb0)
        layout.addWidget(in_name)
        layout.addWidget(lb1)
        layout.addWidget(in_bz)

        # 创建一个按钮框
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(button_box)
        # 设置按钮的点击事件
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        # 将布局设置给对话框
        dialog.setLayout(layout)
        # 显示对话框并等待用户点击按钮
        result = dialog.exec_()
        # 获取用户输入的文本
        if result == QDialog.Accepted:
            return {'name':in_name.text(),'bz':in_bz.text()}
        else:
            return {}


    def select_xuqiu(self):
        selected_text = self.li_search.selectedText().strip()
        if selected_text!='':
            self.init_tb_xuqiu(selected_text)
    def search_xuqiu(self):
        search_str = self.li_search.text().strip().strip()
        if search_str!='':
            self.init_tb_xuqiu(search_str)
        else:
            self.init_tb_xuqiu()






    def remove_xuqiu(self):
        #删除需求
        try:
            cur_index = self.tb_xuqiu.currentIndex()
            data = self.get_row_content(cur_index)
            id = int(data[0])
            pt = data[xuqiu_header_xqccwz]

            result = QMessageBox.question(self, "提醒", "是否删除内容", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if result == QMessageBox.Yes:
                shutil.move(pt,HSZ_PATH)
                XQ_INFO.delete_by_id(id)
                self.init_tb_xuqiu()

            
        except Exception as e:
            QMessageBox.warning(self,'',e.__str__())

    # xuqiu_dtl_header = ['id','类型', '文件名', '版本', '备注','创建日期','更新日期', '路径']
    def inint_tb_xuqiudtl(self,xq_item):
        files,dirs = get_files_in_directory(xq_item.path)
        if self.ra_banben.isChecked():
            items = JB_INFO.select().where(JB_INFO.xq_id==xq_item).order_by(JB_INFO.name)
        else:
            items = JB_INFO.select().where((JB_INFO.xq_id == xq_item) & (JB_INFO.version==0)).order_by(JB_INFO.name)
        
        # md 文档
        md_items = MD_INFO.select().where(MD_INFO.xq_id==xq_item).order_by(MD_INFO.name)

        model = QStandardItemModel()
        for column, name in enumerate(xuqiu_dtl_header):
            header_item = QStandardItem(name)
            model.setHorizontalHeaderItem(column, header_item)
        ct = self.com_type.currentText()
        datas = []

        search_str = self.dict_search_tmp.get(xq_item.id,'')
        self.li_search_wd.setText(search_str)

        search_str = self.li_search_wd.text()

        md_files_names = [item.name for item in md_items]
        ex_file = [item.name for item in items]
        ex_file += md_files_names

        if ct == '' or ct == '文件夹':
            for file in dirs :
                id = '-999'
                name = file
                desc = ''
                type = '文件夹'
                version='0'
                path = os.path.join(xq_item.path,file)
                ctime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime( os.path.getctime(path)))
                mtime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime( os.path.getmtime(path)))
                datas.append([id,type, name, version, desc,ctime,mtime, path])

        for ln_file in files:
            if ln_file.endswith('.ln'):
                path = os.path.join(xq_item.path, ln_file)
                with open(path,'r',encoding='utf-8') as f:
                    lines = f.readlines()
                    print(lines)
                for line in lines:
                    line = line.strip()
                    if ';' in line:
                        file_path = line.split(';')[0]
                        desc = line.split(';')[1]
                    else:
                        file_path = line
                        desc = ''
                    id = '-888'
                    type = '文档'
                    version = '0'
                    name = file_path.split('\\')[-1]
                    if os.path.exists(file_path):
                        ctime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(os.path.getctime(file_path)))
                        mtime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(os.path.getmtime(file_path)))
                        datas.append([id,type, f'@@@{name}', version, desc,ctime,mtime, file_path])
                    if 'http' in file_path:
                        ctime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
                        mtime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
                        datas.append([id,type, f'@@@{file_path[:50]}', version, desc,ctime,mtime, file_path])
        files.reverse()
        if ct == '' or ct == '文档':
            ## md 文档
            for item in md_items:
                id =str(item.file_id)
                name = item.name
                desc = item.bz
                type = 'MD'
                ctime = item.ctime.strftime("%Y-%m-%d %H:%M:%S")
                mtime = item.tj_time.strftime("%Y-%m-%d %H:%M:%S")
                path = item.path
                version = '0'
                datas.append([id, type, name, version, desc,ctime, mtime,path])
            for file in files :
                if file in ex_file:
                    continue
                if file.endswith('.ln'):
                    continue
                id = '-999'
                name = file
                desc = ''
                type = '文档'
                version='0'
                print(f'@@@@@@@@@@@@@file:{name}')
                if os.path.exists(file):
                    tmp = file.split("\\")
                    name = f'@@@@-{tmp[-1]}'
                    path = file
                else:
                    path = os.path.join(xq_item.path,file)
                ctime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime( os.path.getctime(path)))
                mtime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime( os.path.getmtime(path)))

                datas.append([id,type, name, version, desc,ctime,mtime, path])



        if ct == '' or ct == '脚本':
            guidang_set = set() # 为了去重
            # 归档开关
            xianshi_guidang_switch = self.is_guidang.isChecked() 
            jiaoben_data =[]
            guidang_data =[]

            # 应用脚本引入
            yy_jbs = YY_JB_INFO.select().where(YY_JB_INFO.t_xq_id==xq_item)
            for yyjb in yy_jbs:
                id = str(yyjb.id)
                jb_id = yyjb.sjb_id
                jb_info = JB_INFO.get_by_id(jb_id)
                name = jb_info.name
                desc = yyjb.yy_desc
                type = '脚本引用'
                ctime = jb_info.ctime.strftime("%Y-%m-%d %H:%M:%S")
                if not os.path.exists(path):
                    mtime ='文件不存在'
                else:
                    mtime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime( os.path.getmtime(path)))
                path = jb_info.path
                version = str(jb_info.version)
                jiaoben_data.append([id, type, name, version, desc,ctime, mtime,path])


            for item in items:
                id =str(item.id)
                name = item.name
                desc = item.desc
                type = item.type
                ctime = item.ctime.strftime("%Y-%m-%d %H:%M:%S")
                path = item.path
                gd_name = item.gd_name
                if not gd_name:
                    gd_name = ''
                if not os.path.exists(path):
                    mtime ='文件不存在'
                else:
                    mtime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime( os.path.getmtime(path)))
                version = str(item.version)
                
                
                if xianshi_guidang_switch or gd_name == '':
                    if gd_name != '':
                        name = f'{name}->【{gd_name}】'
                    jiaoben_data.append([id, type, name, version, desc,ctime, mtime,path])
                else:
                    if gd_name not in guidang_set:
                        guidang_data.append(['-777', '归档', gd_name, '0', '归档文件','', '',''])
                        guidang_set.add(gd_name)
        
        


            datas = datas + guidang_data+jiaoben_data
        for item in datas:
            if item[2].startswith('~$'):
                continue
            pdyj = [item[2],item[4].upper()]
            if self.xuqiu_dtl_is_display(search_str,pdyj):
            # if search_str.upper() in item[2].upper() or search_str.strip()=='' or search_str.upper() in item[4].upper():
                row_items = [QStandardItem(i) for i in item]
                model.appendRow(row_items)

        self.tb_xuqiudtl.setModel(model)
        self.tb_xuqiudtl.setStyleSheet(mhsheet)

        # self.tb_xuqiudtl.resizeRowsToContents()
        self.repaint()
    # 需求明细是否展示
    def xuqiu_dtl_is_display(self,search,pdyj):
        #search 为关键字，支持& | 运算，|优先级大
        search = search.upper()
        pdyj = u'\u033F'.join(pdyj)
        search_list = search.split('|')
        for item in  search_list:
            flag = True
            search_and_list = item.split('&')
            for and_yj in search_and_list:
                if and_yj not in pdyj.upper():
                    flag = False
            if flag :
                return True
        return  False
    def init_tb_xuqiu(self,search=None):
        # 显示需求表格
        rwlx =''
        
        com_status = self.com_sear_type.currentText()
        if search:
            gjz_str_ss = search
        else:
            gjz_str_ss = self.li_search.text()
        model = QStandardItemModel()
        gjz_str_l =[]
        zd_list = ZDXQ.split('|')
        gjz_str_l += zd_list
        if gjz_str_ss!='':
            gjz_str_l += gjz_str_ss.split('|')
        xs_ids = []

        all_items = []

        for gjz_str in gjz_str_l:
            gjz_list = gjz_str.split('&')
            items = XQ_INFO.select().order_by(XQ_INFO.mtime.desc())
            if com_status!='':
                items = items.where(XQ_INFO.status==com_status)
            for gjz in gjz_list:
                items = items.where((XQ_INFO.name ** f'%{gjz}%')|(XQ_INFO.xuqiu_type == gjz)|(XQ_INFO.tcr  ** f'%{gjz}%'))
            all_items += list(items)
        # 开发周期内的需求
        cur_date = QDate.currentDate().toString('yyyy-MM-dd')

        # cur_date -7 
        cur_date_7 = QDate.currentDate().addDays(-7).toString('yyyy-MM-dd')
        cur_items = XQ_INFO.select().where(XQ_INFO.start_date<=cur_date,XQ_INFO.end_date>=cur_date_7,XQ_INFO.status=='开发中')
        all_items += list(cur_items)
        
        for column, name in enumerate(xuqiu_header):
                header_item = QStandardItem(name)
                model.setHorizontalHeaderItem(column, header_item) 
           
        for item in all_items[:500]:
            id = str(item.id)
            if id in xs_ids:
                continue
            xs_ids.append(id)
            name = item.name
            path = item.path
            xuqiu_type = item.xuqiu_type
            status = item.status
            tcr = item.tcr
            ctime = item.ctime.strftime("%Y-%m-%d %H:%M:%S")
            start_date = item.start_date.strftime("%Y-%m-%d") if item.start_date else ''
            end_date = item.end_date.strftime("%Y-%m-%d") if item.end_date else ''
            rentian = str(item.rentian) if item.rentian else ''
            # 剩余人天, 结束时间减去当前时间
            # 获取当前时间
            if end_date != '' and status == '开发中':
                # 获取当前日期字符串
                cur_date_str = QDate.currentDate().toString('yyyy-MM-dd')
                # 将字符串转为QDate对象
                end_qdate = QDate.fromString(end_date, 'yyyy-MM-dd')
                cur_qdate = QDate.fromString(cur_date_str, 'yyyy-MM-dd')
                # 计算当前日期到结束日期的天数
                srtr = cur_qdate.daysTo(end_qdate)
                srtr = str(srtr)
                # info日志输出
                print(f"[info] 计算剩余人天: 当前日期={cur_date_str}, 结束日期={end_date}, 剩余人天={srtr}")
            else:
                srtr = ''

            row_items = [QStandardItem(i) for i in [id,name,xuqiu_type,status,tcr,path,ctime,start_date,end_date,rentian,srtr]]
            model.appendRow(row_items)
        self.tb_xuqiu.setModel(model)

        #列宽行高
        self.tb_xuqiu.setEditTriggers(QTableView.NoEditTriggers)
        self.repaint()

    def get_row_content(self,index):
        try:
        # 获取行内容
            row = [index.model().data(index.sibling(index.row(), column), Qt.DisplayRole) for column in
                   range(index.model().columnCount())]
            return row
        except Exception as e:
            raise ValueError('未正确选择行')


    def ddj_xuqiu(self,index):
        datas = self.get_row_content(index)
        path = datas[xuqiu_header_xqccwz]
        os.startfile(path)
    def dj_xuqiu(self,index):
        #刷新需求内容
        datas = self.get_row_content(index)
        id = int(datas[0])
        xq_item = XQ_INFO.get_by_id(id)
        self.f_set_la_xq_info(id)
        self.inint_tb_xuqiudtl(xq_item)

    def f_com_type(self):
        row = self.tb_xuqiu.currentIndex()
        datas = self.get_row_content(row)
        id = int(datas[0])
        self.dict_search_tmp[id]=self.li_search_wd.text()
        self.tb_xuqiu.setFocus()
        self.dj_xuqiu(row)
        self.li_search_wd.setFocus()

    # id = AutoField(verbose_name='脚本编号', primary_key=True)
    # xq_id = ForeignKeyField(XQ_INFO,on_delete='CASCADE')
    # name = CharField(verbose_name='文件名称')
    # desc = CharField(verbose_name='备注')
    # type = CharField(verbose_name='类型') #开发/提交（版本）
    # version = IntegerField(verbose_name='版本号')
    # path = CharField(verbose_name='路径')
    def create_xuqiu(self):
        # 创建需求
        if self.li_search.text() == '#*#*脚本迁移*#*#':
            print('开始迁移')
            qylist = JB_INFO.select(JB_INFO,XQ_INFO).join(XQ_INFO).where(JB_INFO.version==0)
            for i in qylist:
                try:
                    xq_id = i.xq_id
                    name = i.name
                    desc = i.desc
                    type = i.type
                    version = 0
                    path = os.path.join(xq_id.path,name)
                    old_jb = JB_INFO.get_by_id(i.id)
                    new_jb = JB_INFO(name=name,desc = desc,type=type,version=version,path = path,xq_id=xq_id)
                    print(old_jb.name,old_jb.path,new_jb.path,new_jb.name)

                    shutil.copy(old_jb.path,new_jb.path)
                    old_jb.version = '-1'
                    new_jb.save()
                    old_jb.save()
                except Exception as e:
                    print(e.__str__())
        #*#*导入表*#*#$$E:\思特奇\SQL_HELPER目录\数据\图书馆\元数据\YSJ (1)\表元数据.xlsx
        if '#*#*导入表*#*#$$' in self.li_search.text() :
            my_path = self.li_search.text()
            my_path = my_path.replace('#*#*导入表*#*#$$','')
            from openpyxl import load_workbook
            wb = load_workbook(filename=my_path)
            ws = wb['Sheet1']
            for  row in ws.iter_rows():
                db = row[0].value.upper().strip()
                tb_name = row[1].value.upper().strip()
                table_comment = f'{row[2].value}-批量导入'
                i = f'{db}.{tb_name}'
                print(i,db, tb_name, table_comment)
                u, is_create = TABLE_INFO.get_or_create(table_id=i)
                if not is_create:
                    continue
                else:
                    if 'TB_MID' in i:
                        u.by1 = '3'
                    elif 'PMID' in i:
                        u.by1 = '2'
                    elif 'PDATA' in i:
                        u.by1 = '1'
                    else:
                        u.by1 = '0'
                    u.desc = table_comment
                    u.save()
        #*#*导入字段*#*#$$E:\思特奇\SQL_HELPER目录\数据\图书馆\元数据\YSJ (1)\字段元数据.xlsx
        if '#*#*导入字段*#*#$$' in self.li_search.text():
            my_path = self.li_search.text()
            my_path = my_path.replace('#*#*导入字段*#*#$$', '')
            from openpyxl import load_workbook
            wb = load_workbook(filename=my_path)
            ws = wb['Sheet1']
            for row in ws.iter_rows():
                db = row[0].value.upper().strip()
                tb_name = row[1].value.upper().strip()
                cl_name = row[2].value.upper().strip()
                cl_type = row[3].value.upper().strip()
                cl_comment = f'{row[4].value}'
                i = f'${db}.{tb_name}.${cl_name}'
                tb_id = f'{db}.{tb_name}'
                print(i, db, tb_name, cl_name,cl_type,cl_comment,tb_id)
                u, is_create = TABLE_COLUM_INFO.get_or_create(id=i,tb_id=tb_id)
                if not is_create:
                    continue
                else:
                    u.col_name = cl_name
                    u.col_type = cl_type
                    u.col_desc = cl_comment
                    u.ex_text ='批量导入'
                    u.save()


        else:
            self.creat_xq = Create_XUQIU_UI(self)
            self.creat_xq.setWindowFlag(QtCore.Qt.WindowStaysOnTopHint)
            self.creat_xq.show()
            self.init_tb_xuqiu()
            self.repaint()


import sys
if __name__ == '__main__':
    app = QApplication(sys.argv)
    xuqiu_ui = XUQIU_UI(None)
    xuqiu_ui.show()
    sys.exit(app.exec_())